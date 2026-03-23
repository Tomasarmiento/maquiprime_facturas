from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

RFC_RECEPTOR_ESPERADO = "MES2301274X9"
TARGET_YEAR = 2026

MESES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

MESES_INV = {v: k.capitalize() for k, v in MESES.items()}

YELLOW = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
RED = PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid")

COLUMNS = [
    "Fecha",
    "Proveedor",
    "Proveedor RFC",
    "Folio Factura",
    "UUID",
    "Concepto",
    "Importe",
    "IVA",
    "Otros Impuestos",
    "Total",
    "Comentarios",
    "Empleado",
]

CATEGORY_EXACT = {
    "15101514": "GASOLINA",
    "15101515": "GASOLINA",
    "95111602": "PEAJE",
    "95111603": "PEAJE",
    "78111807": "ESTACIONAMIENTO",
    "90111800": "HOTEL",
    "90111500": "HOTEL / HOSPEDAJE",
    "90101500": "ALIMENTO / BEBIDA",
    "90101501": "ALIMENTO / BEBIDA",
    "90101503": "ALIMENTO / BEBIDA",
    "90101700": "ALIMENTO / BEBIDA",
    "90101800": "ALIMENTO / BEBIDA",
    "78111804": "TAXI",
    "78111800": "TRANSPORTE",
    "78111808": "ALQUILER DE AUTO",
    "78111811": "ALQUILER DE AUTO",
    "83111603": "DATOS MÓVILES",
    "43201415": "DATOS MÓVILES",
    "27113300": "HERRAMIENTAS",
    "27131500": "HERRAMIENTAS",
    "23291900": "HERRAMIENTAS INDUSTRIALES",
    "14111828": "PAPELERÍA",
}

CATEGORY_PREFIX = {
    "831116": "DATOS MÓVILES",
    "2711": "HERRAMIENTAS",
    "2713151": "HERRAMIENTAS",
    "141115": "PAPELERÍA",
    "441217": "PAPELERÍA",
}


@dataclass
class InvoiceRow:
    fecha: datetime
    proveedor: str
    proveedor_rfc: str
    folio_factura: str
    uuid: str
    concepto: str
    importe: Decimal
    iva: Decimal
    otros_impuestos: Decimal
    total: Decimal
    comentarios: str
    empleado: str
    source_month: str
    source_file: Path


class Processor:
    def __init__(self, base_2026: Path, excel_path: Path, logger):
        self.base_2026 = base_2026
        self.excel_path = excel_path
        self.logger = logger

    @staticmethod
    def _normalize_uuid(value) -> str:
        return str(value or "").strip().upper().replace("{", "").replace("}", "")

    # ------------------------------------------------------------------
    # Main entry point
    # ------------------------------------------------------------------

    def run(self, dry_run: bool = False) -> dict:
        wb = load_workbook(self.excel_path)

        # Step 1: collect UUIDs already present in the workbook
        existing_uuid_positions = self._collect_uuid_positions(wb)
        existing_uuid_set = set(existing_uuid_positions.keys())

        # Step 2: track UUIDs inserted in this run
        new_uuid_positions: dict[str, list[tuple[str, int]]] = {}

        inserted = 0
        warnings = 0
        errors = 0

        # Step 3: iterate month folders → employee folders → XML files
        for month_name, month_num in MESES.items():
            month_folder = self.base_2026 / month_name.capitalize()
            if not month_folder.exists() or not month_folder.is_dir():
                continue

            employee_folders = sorted(
                [p for p in month_folder.iterdir() if p.is_dir()],
                key=lambda p: p.name.lower(),
            )

            for employee_folder in employee_folders:
                for xml_file in sorted(employee_folder.glob("*.xml")):
                    # --- parse ---
                    try:
                        row = self._parse_invoice(xml_file, employee_folder.name, month_name)
                    except Exception as exc:
                        errors += 1
                        self.logger(f"ERROR parseando {xml_file.name}: {exc}")
                        continue

                    # --- validate year ---
                    if row.fecha.year != TARGET_YEAR:
                        errors += 1
                        self.logger(
                            f"ERROR fecha fuera de {TARGET_YEAR} ({row.fecha.date()}) en {xml_file.name}"
                        )
                        continue

                    # --- validate UUID ---
                    if not row.uuid:
                        errors += 1
                        self.logger(f"ERROR UUID vacío en {xml_file.name}")
                        continue

                    uuid_key = self._normalize_uuid(row.uuid)

                    # --- skip if already in workbook ---
                    if uuid_key in existing_uuid_set:
                        self.logger(
                            f"OMITIDO UUID ya existente: {row.uuid} ({xml_file.name})"
                        )
                        continue

                    # --- determine target sheet ---
                    target_sheet = f"{MESES_INV[row.fecha.month]} {TARGET_YEAR}"
                    if target_sheet not in wb.sheetnames:
                        ws = wb.create_sheet(target_sheet)
                        self._ensure_headers(ws)

                    ws = wb[target_sheet]
                    self._ensure_headers(ws)

                    # --- insert row ---
                    if not dry_run:
                        # Insert after the last row that has a UUID, so that
                        # manual rows (sin comprobante, tickets, etc.) always
                        # stay below the program-generated block.
                        self._ensure_separator(ws)
                        insert_at = self._find_insert_position(ws) + 1
                        ws.insert_rows(insert_at)
                        data_to_write = [
                            row.fecha,
                            row.proveedor,
                            row.proveedor_rfc,
                            row.folio_factura,
                            row.uuid,
                            row.concepto,
                            float(row.importe),
                            float(row.iva),
                            float(row.otros_impuestos),
                            None,  # Total: written as formula below
                            row.comentarios,
                            row.empleado,
                        ]
                        for col_idx, val in enumerate(data_to_write, start=1):
                            ws.cell(row=insert_at, column=col_idx, value=val)
                        # Write Total as formula
                        ws.cell(row=insert_at, column=10).value = f"=G{insert_at}+H{insert_at}+I{insert_at}"
                        inserted_row = insert_at
                        # Shift existing uuid positions down by 1 for this sheet
                        for key, positions in existing_uuid_positions.items():
                            existing_uuid_positions[key] = [
                                (s, r + 1) if s == ws.title and r >= insert_at else (s, r)
                                for s, r in positions
                            ]
                        for key, positions in new_uuid_positions.items():
                            new_uuid_positions[key] = [
                                (s, r + 1) if s == ws.title and r >= insert_at else (s, r)
                                for s, r in positions
                            ]
                        inserted += 1

                        # yellow highlight if invoice date doesn't match folder month
                        if row.source_month.lower() != MESES_INV[row.fecha.month].lower():
                            warnings += 1
                            self._fill_row(ws, inserted_row, YELLOW)
                            self.logger(
                                f"ADVERTENCIA mes distinto: {xml_file.name} "
                                f"en carpeta '{row.source_month}' pero fecha {row.fecha.date()}"
                            )

                        # track for duplicate detection
                        new_uuid_positions.setdefault(uuid_key, []).append(
                            (ws.title, inserted_row)
                        )
                        # mark as seen so we don't insert the same UUID twice
                        # even if two XML files in this run share the same UUID
                        existing_uuid_set.add(uuid_key)

        # Step 4: apply red highlight to any duplicated UUIDs
        if not dry_run:
            dup_count = self._apply_duplicates(wb, existing_uuid_positions, new_uuid_positions)
            warnings += dup_count
            if dup_count:
                self.logger(f"ADVERTENCIA UUIDs duplicados detectados y marcados en rojo: {dup_count}")

            # Step 5: sort only newly inserted rows at bottom of each sheet
            # (existing rows and their manual edits/colors are never touched)
            try:
                self._sort_new_rows(wb, new_uuid_positions)
            except Exception as exc:
                errors += 1
                self.logger(f"ERROR ordenando filas nuevas: {exc}")

            # Refresh Total formula on every UUID row in every sheet
            self._refresh_total_formulas(wb)
            wb.save(self.excel_path)

        return {
            "inserted": inserted,
            "warnings": warnings,
            "errors": errors,
            "dry_run": dry_run,
        }

    # ------------------------------------------------------------------
    # UUID helpers
    # ------------------------------------------------------------------

    def _collect_uuid_positions(self, wb) -> dict[str, list[tuple[str, int]]]:
        """Return a mapping of normalised UUID → [(sheet_name, row_number)] for every
        data row already in the workbook."""
        result: dict[str, list[tuple[str, int]]] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.max_row < 2:
                continue
            headers = [ws.cell(1, i + 1).value for i in range(len(COLUMNS))]
            if headers[: len(COLUMNS)] != COLUMNS:
                continue
            for r in range(2, ws.max_row + 1):
                raw = ws.cell(r, 5).value
                if raw:
                    key = self._normalize_uuid(raw)
                    if key:
                        result.setdefault(key, []).append((name, r))
        return result

    def _apply_duplicates(
        self,
        wb,
        existing: dict[str, list[tuple[str, int]]],
        new: dict[str, list[tuple[str, int]]],
    ) -> int:
        """Highlight in red every row (old + new) that shares a UUID with another row."""
        dup_count = 0
        for uuid_key, new_positions in new.items():
            all_positions = list(existing.get(uuid_key, [])) + new_positions
            if len(all_positions) > 1:
                dup_count += 1
                for sheet_name, row_num in all_positions:
                    if sheet_name in wb.sheetnames:
                        self._fill_row(wb[sheet_name], row_num, RED)
        return dup_count

    # ------------------------------------------------------------------
    # Excel helpers
    # ------------------------------------------------------------------

    def _fill_row(self, ws, row_number: int, fill: PatternFill) -> None:
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(row=row_number, column=c).fill = copy(fill)

    def _ensure_headers(self, ws) -> None:
        if ws.cell(1, 1).value != COLUMNS[0]:
            for idx, col in enumerate(COLUMNS, start=1):
                ws.cell(row=1, column=idx, value=col)
        # Always ensure autofilter covers the header row
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
        # Remove ALL empty rows in the sheet (not just at the top).
        # openpyxl reports rows with only formatting (no value) as non-empty,
        # so we delete rows where every cell has no value regardless of format.
        rows_to_delete = []
        for r in range(2, ws.max_row + 1):
            if all(ws.cell(r, c).value is None for c in range(1, len(COLUMNS) + 1)):
                rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            ws.delete_rows(r)

    def _detect_row_highlight(self, ws, row_number: int) -> str | None:
        rgb = (
            getattr(getattr(ws.cell(row=row_number, column=1).fill, "start_color", None), "rgb", "")
            or ""
        ).upper()
        if "EF9A9A" in rgb:
            return "RED"
        if "FFF59D" in rgb:
            return "YELLOW"
        return None

    def _sort_new_rows(
        self,
        wb,
        new_uuid_positions: dict[str, list[tuple[str, int]]],
    ) -> None:
        """Sort only the newly added rows within each sheet, leaving existing rows untouched.

        For each sheet that received new rows this run, the new rows are read,
        sorted by (employee, date), then rewritten in place — the existing rows
        above them are never modified.
        """
        # Build per-sheet list of new row numbers
        sheet_new_rows: dict[str, list[int]] = {}
        for positions in new_uuid_positions.values():
            for sheet_name, row_num in positions:
                sheet_new_rows.setdefault(sheet_name, []).append(row_num)

        for sheet_name, row_numbers in sheet_new_rows.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            row_numbers_sorted = sorted(row_numbers)

            # Read only the new rows
            data: list[tuple[str, datetime, list, str | None]] = []
            for r in row_numbers_sorted:
                values = [ws.cell(r, c).value for c in range(1, len(COLUMNS) + 1)]
                highlight = self._detect_row_highlight(ws, r)
                empleado = str(values[11] or "").lower()
                fecha_dt = self._coerce_datetime(values[0], sheet_name, r)
                if fecha_dt is None:
                    fecha_dt = datetime.max
                data.append((empleado, fecha_dt, values, highlight))

            data.sort(key=lambda x: (x[0], x[1]))

            # Rewrite only those rows
            for r, (_, _, values, highlight) in zip(row_numbers_sorted, data):
                for c in range(1, len(COLUMNS) + 1):
                    ws.cell(r, c).value = values[c - 1]
                if highlight == "YELLOW":
                    self._fill_row(ws, r, YELLOW)
                elif highlight == "RED":
                    self._fill_row(ws, r, RED)
                else:
                    # clear fill only on new rows (safe: user hasn't touched them yet)
                    for c in range(1, len(COLUMNS) + 1):
                        ws.cell(r, c).fill = PatternFill(fill_type=None)


    def _refresh_total_formulas(self, wb) -> None:
        """Rewrite =G{n}+H{n}+I{n} in column J for every data row in every sheet
        (both UUID rows and manual rows like Sin comprobante/Ticket).
        This corrects any formula drift caused by insert_rows shifting rows."""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue
            for r in range(2, ws.max_row + 1):
                # Any row that has data in col 1 or col 2 (fecha or proveedor)
                has_data = ws.cell(r, 1).value or ws.cell(r, 2).value
                is_separator = ws.cell(r, 2).value == self.SEPARATOR
                if has_data and not is_separator:
                    ws.cell(r, 10).value = f"=G{r}+H{r}+I{r}"

    SEPARATOR = "--- FIN FACTURAS ---"

    def _find_insert_position(self, ws) -> int:
        """Return the row number just before the separator row.
        If no separator exists, insert after the last UUID row.
        If no UUID rows exist, insert at row 2."""
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value == self.SEPARATOR:
                return r - 1  # insert before separator
        # No separator found — fall back to after last UUID row
        last = 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 5).value:
                last = r
        return last

    def _ensure_separator(self, ws) -> None:
        """Make sure the separator row exists. If not, add it after the last UUID row."""
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value == self.SEPARATOR:
                return  # already there
        # Add it after last UUID row
        last_uuid = 1
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 5).value:
                last_uuid = r
        sep_row = last_uuid + 1
        ws.insert_rows(sep_row)
        ws.cell(sep_row, 2).value = self.SEPARATOR
        from openpyxl.styles import Font, PatternFill, Alignment
        sep_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        sep_font = Font(bold=True, color="666666", italic=True)
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(sep_row, c).fill = sep_fill
            ws.cell(sep_row, c).font = sep_font
        ws.cell(sep_row, 2).alignment = Alignment(horizontal="center")

    def _coerce_datetime(self, value, sheet_name: str, row_number: int) -> datetime | None:
        if isinstance(value, datetime):
            return value
        if value in (None, ""):
            self.logger(
                f"ADVERTENCIA fecha vacía en sheet '{sheet_name}' fila {row_number}; se ordena al final."
            )
            return None
        try:
            return datetime.fromisoformat(str(value))
        except ValueError:
            self.logger(
                f"ADVERTENCIA fecha inválida '{value}' en sheet '{sheet_name}' fila {row_number}; se ordena al final."
            )
            return None

    # ------------------------------------------------------------------
    # XML parsing
    # ------------------------------------------------------------------

    def _parse_invoice(self, xml_file: Path, empleado: str, source_month: str) -> InvoiceRow:
        doc = etree.parse(str(xml_file))
        root = doc.getroot()
        ns = {
            "cfdi": "http://www.sat.gob.mx/cfd/4",
            "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital",
        }

        fecha_str = root.get("Fecha", "")
        fecha = datetime.fromisoformat(fecha_str.replace("Z", "+00:00"))

        emisor = root.find("cfdi:Emisor", namespaces=ns)
        receptor = root.find("cfdi:Receptor", namespaces=ns)

        proveedor = emisor.get("Nombre", "") if emisor is not None else ""
        proveedor_rfc = emisor.get("Rfc", "") if emisor is not None else ""

        receptor_rfc = receptor.get("Rfc", "") if receptor is not None else ""
        if receptor_rfc != RFC_RECEPTOR_ESPERADO:
            raise ValueError(f"RFC receptor inválido: {receptor_rfc}")

        uuid = ""
        complemento = root.find("cfdi:Complemento", namespaces=ns)
        if complemento is not None:
            tfd_node = complemento.find("tfd:TimbreFiscalDigital", namespaces=ns)
            if tfd_node is not None:
                uuid = self._normalize_uuid(tfd_node.get("UUID", ""))

        serie = root.get("Serie", "")
        folio = root.get("Folio", "")
        folio_factura = f"{serie}-{folio}" if serie and folio else (folio or serie)

        subtotal = Decimal(root.get("SubTotal", "0"))
        total = Decimal(root.get("Total", "0"))

        iva = Decimal("0")
        otros = Decimal("0")
        impuestos = root.find("cfdi:Impuestos", namespaces=ns)
        if impuestos is not None:
            for tax in impuestos.findall("cfdi:Traslados/cfdi:Traslado", namespaces=ns):
                amount = Decimal(tax.get("Importe", "0"))
                if tax.get("Impuesto", "") == "002":
                    iva += amount
                else:
                    otros += amount
            for tax in impuestos.findall("cfdi:Retenciones/cfdi:Retencion", namespaces=ns):
                amount = Decimal(tax.get("Importe", "0"))
                if tax.get("Impuesto", "") == "002":
                    iva += amount
                else:
                    otros += amount

        concepto_code = self._first_concept_code(root, ns)
        concepto = self._map_category(concepto_code)

        return InvoiceRow(
            fecha=fecha,
            proveedor=proveedor,
            proveedor_rfc=proveedor_rfc,
            folio_factura=folio_factura,
            uuid=uuid,
            concepto=concepto,
            importe=subtotal,
            iva=iva,
            otros_impuestos=otros,
            total=total,
            comentarios="",
            empleado=empleado,
            source_month=source_month,
            source_file=xml_file,
        )

    def _first_concept_code(self, root, ns) -> str:
        conceptos = root.findall("cfdi:Conceptos/cfdi:Concepto", namespaces=ns)
        return conceptos[0].get("ClaveProdServ", "") if conceptos else ""

    def _map_category(self, clave: str) -> str:
        if not clave:
            return ""
        if clave in CATEGORY_EXACT:
            return CATEGORY_EXACT[clave]
        for prefix, category in CATEGORY_PREFIX.items():
            if clave.startswith(prefix):
                return category
        return clave


def discover_month_folders(base_2026: Path) -> Iterable[Path]:
    for p in sorted(base_2026.iterdir(), key=lambda x: x.name.lower()):
        if p.is_dir() and p.name.lower() in MESES:
            yield p