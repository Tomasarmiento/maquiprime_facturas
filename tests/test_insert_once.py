import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from processor import COLUMNS, Processor


def _cfdi_xml(uuid: str, fecha: str = "2026-01-13T14:13:12") -> str:
    return f'''<?xml version="1.0" encoding="utf-8"?>
<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4" xmlns:tfd="http://www.sat.gob.mx/TimbreFiscalDigital" Fecha="{fecha}" SubTotal="100.00" Total="116.00" Serie="A" Folio="1">
  <cfdi:Emisor Rfc="OEMF830516FD0" Nombre="PROVEEDOR TEST"/>
  <cfdi:Receptor Rfc="MES2301274X9" Nombre="MAQUIPRIME"/>
  <cfdi:Conceptos>
    <cfdi:Concepto ClaveProdServ="15101515" Importe="100.00"/>
  </cfdi:Conceptos>
  <cfdi:Impuestos>
    <cfdi:Traslados>
      <cfdi:Traslado Impuesto="002" Importe="16.00"/>
    </cfdi:Traslados>
  </cfdi:Impuestos>
  <cfdi:Complemento>
    <tfd:TimbreFiscalDigital UUID="{uuid}"/>
  </cfdi:Complemento>
</cfdi:Comprobante>
'''


class InsertOnceTest(unittest.TestCase):
    def test_existing_uuid_is_not_reinserted(self):
        with TemporaryDirectory() as td:
            base = Path(td) / "2026"
            (base / "Enero" / "David").mkdir(parents=True)
            excel = base / "FICHERO_CONTROL_2026.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Enero 2026"
            for i, h in enumerate(COLUMNS, start=1):
                ws.cell(1, i, h)
            # Existing UUID in lower-case with braces to test normalization.
            ws.append([
                "2026-01-13 14:13:12",
                "PROVEEDOR TEST",
                "OEMF830516FD0",
                "A-1",
                "{72dceac0-673a-4880-919d-fab941eb398a}",
                "GASOLINA",
                100,
                16,
                0,
                116,
                "",
                "David",
            ])
            wb.save(excel)

            xml_uuid = "72DCEAC0-673A-4880-919D-FAB941EB398A"
            (base / "Enero" / "David" / "factura1.xml").write_text(_cfdi_xml(xml_uuid), encoding="utf-8")

            logs = []
            result = Processor(base, excel, logs.append).run(dry_run=False)
            self.assertEqual(result["inserted"], 0)

            wb2 = load_workbook(excel)
            ws2 = wb2["Enero 2026"]
            self.assertEqual(ws2.max_row, 2)
            self.assertTrue(any("se omite inserción" in m for m in logs))


if __name__ == "__main__":
    unittest.main()
